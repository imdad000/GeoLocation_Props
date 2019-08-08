from django.shortcuts import render
import openpyxl
import pandas as pd
import requests
import xlwt 
from xlwt import Workbook
from django.http import HttpResponse
api_key='AIzaSyDth5Uo5VxsRJVFncZBZToKV1sYLviv6Iw'

def index(request):
    if "GET" == request.method:
        return render(request, 'MyApp/home.html', {})
    else:
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="YourResult.xls"'
        excel_file = request.FILES["excel_file"]


        df = pd.read_excel(excel_file, sheet_name='Sheet1')
    

        column_names=list(df.columns.values)

        df['Latitude']=''
        df['Longitude']=''
        

        for index, row in df.iterrows():
            string=''
            for r in column_names:
                #print(r)
                string=string +str(row[r])+' '
            geocode_url = "https://maps.googleapis.com/maps/api/geocode/json?address={}".format(string)
            print(string)
            if api_key is not None:
                geocode_url = geocode_url + "&key={}".format(api_key)
                results = requests.get(geocode_url)
                results = results.json()
                if len(results['results']) == 0:
                    df['Latitude']=None
                    df['Longitude']=None
                    print("NOT Found")
                else:
                    answer = results['results'][0]
                    lat=answer.get('geometry').get('location').get('lat')
                    lng=answer.get('geometry').get('location').get('lng')
                    print(lat)
                    print(lng)
                    df['Latitude']=lat
                    df['Longitude']=lng
        print(df)
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        worksheet = wb["Sheet1"]
        active_sheet = wb.active
        excel_data = list()
        wb = Workbook() 
        sheet1 = wb.add_sheet('Sheet1')

        i=0
        j=0
        for index, row in df.iterrows():
            row_data = list()
            j=0
            for cell in row:
                sheet1.write(i,j,str(cell))
                row_data.append(str(cell))
                j=j+1
                print(j)
                #print(cell.value)
            excel_data.append(row_data)
            i=i+1


        wb.save(response)
        return response